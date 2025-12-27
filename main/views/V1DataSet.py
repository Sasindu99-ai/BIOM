from drf_spectacular.utils import extend_schema

from vvecon.zorion.auth import Authorized
from vvecon.zorion.logger import Logger
from vvecon.zorion.serializers import Return
from vvecon.zorion.views import API, DeleteMapping, GetMapping, Mapping, PostMapping, PutMapping

from ..payload.requests import FilterDataSetRequest, StudyVariableRequest
from ..payload.responses import DataSetResponse, StudyVariableResponse
from ..services import PatientService, StudyService

__all__ = ['V1DataSet']


@Mapping('api/v1/dataset')
class V1DataSet(API):
	studyService: StudyService = StudyService()
	patientService: PatientService = PatientService()

	# Column patterns to skip when auto-creating variables (from patient match output)
	SKIP_COLUMN_PATTERNS = (
		'matched_',
		'file_duplicate_',
		'file_patient_group',
		'match_status',
		'match_confidence',
		'row_number',
	)

	# Patterns for auto-selecting patient columns
	PATIENT_COLUMN_PATTERNS = {
		'reference': ['patientreference', 'reference', 'ref', 'patient_ref', 'patientid', 'patient_id'],
		'firstName': ['firstname', 'first_name', 'fname', 'first', 'givenname', 'given_name'],
		'lastName': ['lastname', 'last_name', 'lname', 'last', 'surname', 'familyname', 'family_name'],
		'dateOfBirth': ['dob', 'dateofbirth', 'date_of_birth', 'birthdate', 'birth_date', 'birthday'],
		'age': ['age', 'patient_age', 'patientage'],
		'gender': ['gender', 'sex', 'patient_gender', 'patientgender'],
		'latitude': ['latitude', 'lat', 'gps_lat', 'gpslat'],
		'longitude': ['longitude', 'lng', 'lon', 'gps_lng', 'gpslon', 'gps_lon'],
	}

	@extend_schema(
		tags=['Dataset'],
		summary='Get datasets',
		description='Get datasets with filtering, search, and pagination',
		request=FilterDataSetRequest,
		responses={200: DataSetResponse().response()},
	)
	@PostMapping('/')
	@Authorized(True, permissions=['main.view_study'])
	def getDatasets(self, request, data: FilterDataSetRequest):
		Logger.info(f'Validating dataset filter data: {data.initial_data}')
		if data.is_valid(raise_exception=True):
			Logger.info('Dataset filter data is valid')

			# Get pagination info before filtering
			validated_data = data.validated_data.copy()
			pagination_data = validated_data.get('pagination', {})
			page = pagination_data.get('page', 1) if pagination_data else 1
			limit = pagination_data.get('limit', 20) if pagination_data else 20

			# Get filtered queryset WITHOUT pagination for stats
			search_data = validated_data.copy()
			search_data.pop('pagination', None)  # Remove pagination to get all filtered results
			filtered_queryset = self.studyService.search(search_data)

			# Apply sorting
			sort_field = validated_data.get('sortField', 'created_at')
			sort_direction = validated_data.get('sortDirection', 'desc')

			# Map frontend field names to model field names
			field_mapping = {
				'name': 'name',
				'created': 'created_at',
				'version': 'version',
				'category': 'category',
				'created_at': 'created_at',
			}

			# Get the actual field name
			actual_field = field_mapping.get(sort_field, 'created_at')

			# Apply ordering
			if sort_direction == 'asc':
				filtered_queryset = filtered_queryset.order_by(actual_field)
			else:
				filtered_queryset = filtered_queryset.order_by(f'-{actual_field}')

			# Calculate statistics from filtered results
			from django.db.models import Count, Max

			total_count = filtered_queryset.count()

			# Get aggregated stats
			stats_query = filtered_queryset.aggregate(
				totalVariables=Count('variables', distinct=True),
				totalUserStudies=Count('userStudies', distinct=True),
				maxVersion=Max('version'),
			)

			# Now get paginated datasets using paginate method
			datasets = self.studyService.paginate(filtered_queryset, page, limit)

			# Build response
			total_pages = (total_count + limit - 1) // limit if limit > 0 else 1

			response_data = {
				'results': DataSetResponse(data=datasets, many=True).json().data,
				'pagination': {
					'page': page,
					'limit': limit,
					'total': total_count,
					'totalPages': total_pages,
					'hasNext': page < total_pages,
					'hasPrev': page > 1,
				},
				'stats': {
					'total': total_count,
					'totalVariables': stats_query['totalVariables'] or 0,
					'totalUserStudies': stats_query['totalUserStudies'] or 0,
					'latestVersion': stats_query['maxVersion'] or 1,
				},
			}

			Logger.info(f'{len(datasets)} datasets found on page {page}/{total_pages}')
			return Return.ok(response_data)

	@extend_schema(
		tags=['Dataset'],
		summary='Get dataset',
		description='Get single dataset details',
		responses={200: DataSetResponse().response()},
	)
	@GetMapping('/<int:id>')
	@Authorized(True, permissions=['main.view_study'])
	def getDataset(self, request, id: int):
		Logger.info(f'Fetching dataset {id}')
		dataset = self.studyService.getById(id)
		Logger.info(f'Dataset {dataset} fetched')
		return DataSetResponse(data=dataset).json()

	@extend_schema(
		tags=['Dataset'],
		summary='Delete dataset',
		description='Delete dataset',
	)
	@DeleteMapping('/<int:id>')
	@Authorized(True, permissions=['main.delete_study'])
	def deleteDataset(self, request, id: int):
		Logger.info(f'Deleting dataset {id}')
		self.studyService.delete(id)
		Logger.info(f'Dataset {id} deleted')
		return Return.ok()

	# ========== Phase 2 Endpoints ==========

	@extend_schema(
		tags=['Dataset'],
		summary='Get dataset details',
		description='Get full dataset details with variables and statistics',
	)
	@GetMapping('/<int:id>/details')
	@Authorized(True, permissions=['main.view_study'])
	def getDatasetDetails(self, request, id: int):
		Logger.info(f'Fetching dataset details for {id}')
		details = self.studyService.getDetails(id)

		response_data = {
			'dataset': DataSetResponse(data=details['study']).json().data,
			'variables': StudyVariableResponse(data=details['variables'], many=True).json().data,
			'stats': details['stats'],
		}

		Logger.info(f'Dataset {id} details fetched')
		return Return.ok(response_data)

	@extend_schema(
		tags=['Dataset'],
		summary='Get dataset variables',
		description='Get all variables for a dataset',
	)
	@GetMapping('/<int:id>/variables')
	@Authorized(True, permissions=['main.view_study'])
	def getDatasetVariables(self, request, id: int):
		Logger.info(f'Fetching variables for dataset {id}')
		variables = self.studyService.getVariables(id)
		Logger.info(f'{len(variables)} variables found for dataset {id}')
		return StudyVariableResponse(data=variables, many=True).json()

	@extend_schema(
		tags=['Dataset'],
		summary='Add variable to dataset',
		description='Add a new variable to a dataset',
		request=StudyVariableRequest,
	)
	@PostMapping('/<int:id>/variables')
	@Authorized(True, permissions=['main.change_study'])
	def addDatasetVariable(self, request, id: int, data: StudyVariableRequest):
		Logger.info(f'Adding variable to dataset {id}')
		if data.is_valid(raise_exception=True):
			variable = self.studyService.addVariable(id, data.validated_data)
			Logger.info(f'Variable {variable.id} added to dataset {id}')
			return StudyVariableResponse(data=variable).json()

	@extend_schema(
		tags=['Dataset'],
		summary='Update variable',
		description='Update an existing variable',
		request=StudyVariableRequest,
	)
	@PutMapping('/<int:dataset_id>/variables/<int:variable_id>')
	@Authorized(True, permissions=['main.change_study'])
	def updateDatasetVariable(self, request, dataset_id: int, variable_id: int, data: StudyVariableRequest):
		Logger.info(f'Updating variable {variable_id} in dataset {dataset_id}')
		if data.is_valid(raise_exception=True):
			variable = self.studyService.updateVariable(variable_id, data.validated_data)
			Logger.info(f'Variable {variable_id} updated')
			return StudyVariableResponse(data=variable).json()

	@extend_schema(
		tags=['Dataset'],
		summary='Delete variable',
		description='Remove a variable from a dataset',
	)
	@DeleteMapping('/<int:dataset_id>/variables/<int:variable_id>')
	@Authorized(True, permissions=['main.change_study'])
	def deleteDatasetVariable(self, request, dataset_id: int, variable_id: int):
		Logger.info(f'Removing variable {variable_id} from dataset {dataset_id}')
		self.studyService.removeVariable(dataset_id, variable_id)
		Logger.info(f'Variable {variable_id} removed from dataset {dataset_id}')
		return Return.ok()

	@extend_schema(
		tags=['Dataset'],
		summary='Get data preview',
		description='Get paginated data preview for a dataset',
	)
	@PostMapping('/<int:id>/data')
	@Authorized(True, permissions=['main.view_study'])
	def getDataPreview(self, request, id: int):
		Logger.info(f'Fetching data preview for dataset {id}')

		# Get pagination from request body
		page = request.data.get('page', 1)
		limit = request.data.get('limit', 10)

		data_preview = self.studyService.getDataPreview(id, page, limit)
		Logger.info(f'{len(data_preview["rows"])} data rows found for dataset {id}')
		return Return.ok(data_preview)

	@extend_schema(
		tags=['Dataset'],
		summary='Get dataset history',
		description='Get update history timeline for a dataset',
	)
	@GetMapping('/<int:id>/history')
	@Authorized(True, permissions=['main.view_study'])
	def getDatasetHistory(self, request, id: int):
		Logger.info(f'Fetching history for dataset {id}')
		history = self.studyService.getHistory(id)
		Logger.info(f'{len(history)} history events found for dataset {id}')
		return Return.ok({'events': history})

	@extend_schema(
		tags=['Dataset'],
		summary='Get dataset patients',
		description='Get patients who have data entries in this dataset with pagination',
	)
	@GetMapping('/<int:id>/patients')
	@Authorized(True, permissions=['main.view_study'])
	def getDatasetPatients(self, request, id: int):
		Logger.info(f'Fetching patients for dataset {id}')

		page = int(request.GET.get('page', 1))
		limit = int(request.GET.get('limit', 20))

		patients_data = self.studyService.getPatients(id, page, limit)
		Logger.info(f'{len(patients_data.get("patients", []))} patients found for dataset {id}')
		return Return.ok(patients_data)

	@extend_schema(
		tags=['Dataset'],
		summary='Download import template',
		description='Download a template (CSV or Excel) with columns for patient info and dataset variables',
	)
	@GetMapping('/<int:id>/template')
	@Authorized(True, permissions=['main.view_study'])
	def downloadTemplate(self, request, id: int):
		Logger.info(f'Generating import template for dataset {id}')

		import csv
		import io

		from django.http import HttpResponse, StreamingHttpResponse

		# Get format (csv or xlsx) - use 'file_type' param to avoid DRF 'format' conflict
		file_format = request.GET.get('file_type', 'csv').lower()

		# Get dataset and variables
		dataset = self.studyService.getById(id)
		variables = list(dataset.variables.all().order_by('order', 'name'))

		# Define patient info columns (canonical names)
		patient_info_cols = ['PatientReference', 'FirstName', 'LastName', 'DateOfBirth', 'Age', 'Gender', 'Latitude', 'Longitude']

		# Build unique headers: patient info + variables (excluding duplicates)
		used_names = set(col.lower() for col in patient_info_cols)
		headers = patient_info_cols.copy()
		variable_start_index = len(headers)

		for v in variables:
			# Skip if variable name matches a patient column
			if v.name.lower() in used_names:
				continue
			headers.append(v.name)
			used_names.add(v.name.lower())

		safe_name = ''.join(c for c in dataset.name if c.isalnum() or c in (' ', '-', '_')).strip()

		if file_format == 'xlsx':
			# Generate Excel file with data validation
			try:
				from openpyxl import Workbook
				from openpyxl.comments import Comment
				from openpyxl.styles import Alignment, Font, PatternFill
				from openpyxl.utils import get_column_letter
				from openpyxl.worksheet.datavalidation import DataValidation

				wb = Workbook()
				ws = wb.active
				ws.title = 'Import Data'

				# Styles
				patient_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
				variable_fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
				header_font = Font(bold=True)

				# Write headers with styling
				for col_idx, header in enumerate(headers, 1):
					cell = ws.cell(row=1, column=col_idx, value=header)
					cell.font = header_font
					cell.alignment = Alignment(horizontal='center')

					# Color coding
					if col_idx <= variable_start_index:
						cell.fill = patient_fill
					else:
						cell.fill = variable_fill

				# Add column comments (hints)
				hints = {
					'PatientReference': 'Unique patient ID (e.g., PATIENT-001)',
					'FirstName': 'Patient first name',
					'LastName': 'Patient last/family name',
					'DateOfBirth': 'Format: YYYY-MM-DD',
					'Age': 'Number, will create fake DOB if no DOB provided',
					'Gender': 'M or F',
					'Latitude': 'GPS latitude (decimal)',
					'Longitude': 'GPS longitude (decimal)',
				}
				for col_idx, header in enumerate(headers, 1):
					if header in hints:
						ws.cell(row=1, column=col_idx).comment = Comment(hints[header], 'System')

				# Add hints for variables based on type
				for v in variables:
					if v.name.lower() in [col.lower() for col in patient_info_cols]:
						continue
					try:
						col_idx = headers.index(v.name) + 1
						if v.type == 'NUMBER':
							ws.cell(row=1, column=col_idx).comment = Comment('Numeric value', 'System')
						elif v.type == 'DATE':
							ws.cell(row=1, column=col_idx).comment = Comment('Format: YYYY-MM-DD', 'System')
						elif v.type == 'BOOLEAN':
							ws.cell(row=1, column=col_idx).comment = Comment('Yes or No', 'System')
					except ValueError:
						pass

				# Data validations
				# Gender validation
				gender_col = get_column_letter(headers.index('Gender') + 1)
				gender_dv = DataValidation(type='list', formula1='"M,F"', allow_blank=True)
				gender_dv.error = 'Please select M or F'
				gender_dv.prompt = 'Select gender'
				ws.add_data_validation(gender_dv)
				gender_dv.add(f'{gender_col}2:{gender_col}1000')

				# Boolean validations for boolean variables
				for v in variables:
					if v.type == 'BOOLEAN' and v.name in headers:
						col_letter = get_column_letter(headers.index(v.name) + 1)
						bool_dv = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True)
						bool_dv.error = 'Please select Yes or No'
						ws.add_data_validation(bool_dv)
						bool_dv.add(f'{col_letter}2:{col_letter}1000')

				# Add sample rows
				sample_data = [
					['PATIENT-001', 'John', 'Doe', '1985-03-15', '39', 'M', '', ''],
					['PATIENT-002', 'Jane', 'Smith', '', '45', 'F', '', ''],
					['PATIENT-003', '', '', '', '32', '', '6.9271', '79.8612'],
				]
				for row_idx, sample in enumerate(sample_data, 2):
					for col_idx, value in enumerate(sample, 1):
						if col_idx <= len(sample):
							ws.cell(row=row_idx, column=col_idx, value=value)

				# Auto-size columns
				for col_idx, header in enumerate(headers, 1):
					ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(header) + 2)

				# Save to bytes
				output = io.BytesIO()
				wb.save(output)
				output.seek(0)

				response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
				response['Content-Disposition'] = f'attachment; filename="{safe_name}_import_template.xlsx"'

				Logger.info(f'Excel template generated for dataset {id} with {len(variables)} variables')
				return response

			except ImportError:
				Logger.warning('openpyxl not installed, falling back to CSV')
				file_format = 'csv'

		# Generate CSV (default)
		def generate_csv():
			output = io.StringIO()
			writer = csv.writer(output)

			writer.writerow(headers)
			yield output.getvalue()
			output.seek(0)
			output.truncate(0)

			# Sample rows
			sample_data = [
				('PATIENT-001', 'John', 'Doe', '1985-03-15', '39', 'M', '', ''),
				('PATIENT-002', 'Jane', 'Smith', '', '45', 'F', '', ''),
				('PATIENT-003', '', '', '', '32', '', '6.9271', '79.8612'),
			]
			for sample in sample_data:
				row = list(sample)
				# Pad with empty values for variables
				while len(row) < len(headers):
					row.append('')
				writer.writerow(row)
				yield output.getvalue()
				output.seek(0)
				output.truncate(0)

		response = StreamingHttpResponse(generate_csv(), content_type='text/csv')
		response['Content-Disposition'] = f'attachment; filename="{safe_name}_import_template.csv"'

		Logger.info(f'CSV template generated for dataset {id} with {len(variables)} variables')
		return response

	@extend_schema(
		tags=['Dataset'],
		summary='Create dataset',
		description='Create a new dataset',
	)
	@PostMapping('/create')
	@Authorized(True, permissions=['main.add_study'])
	def createDataset(self, request):
		Logger.info('Creating new dataset')
		data = request.data

		# Create the dataset
		from ..models import Study
		dataset = Study.objects.create(
			name=data.get('name'),
			description=data.get('description', ''),
			category=data.get('category'),
			status=data.get('status', 'ACTIVE'),
			reference=data.get('reference', ''),
			createdBy=request.user,
			version=1,
		)

		Logger.info(f'Dataset {dataset.id} created')
		return DataSetResponse(data=dataset).json()

	@extend_schema(
		tags=['Dataset'],
		summary='Update dataset',
		description='Update an existing dataset',
	)
	@PutMapping('/<int:id>')
	@Authorized(True, permissions=['main.change_study'])
	def updateDataset(self, request, id: int):
		Logger.info(f'Updating dataset {id}')
		data = request.data

		dataset = self.studyService.getById(id)

		# Track changes for history
		changes = []
		if 'name' in data and data['name'] != dataset.name:
			changes.append(f"Name changed from '{dataset.name}' to '{data['name']}'")
			dataset.name = data['name']
		if 'description' in data and data['description'] != dataset.description:
			changes.append('Description updated')
			dataset.description = data['description']
		if 'category' in data and data['category'] != dataset.category:
			changes.append(f"Category changed to '{data['category']}'")
			dataset.category = data['category']
		if 'status' in data and data['status'] != dataset.status:
			changes.append(f"Status changed to '{data['status']}'")
			dataset.status = data['status']
		if 'reference' in data and data['reference'] != dataset.reference:
			changes.append('Reference updated')
			dataset.reference = data['reference']

		# Increment version if there are changes
		if changes:
			dataset.version = (dataset.version or 1) + 1
			dataset.save()
			Logger.info(f'Dataset {id} updated to version {dataset.version}: {", ".join(changes)}')
		else:
			Logger.info(f'No changes made to dataset {id}')

		return DataSetResponse(data=dataset).json()

	@extend_schema(
		tags=['Dataset'],
		summary='Preview import data',
		description='Parse uploaded file and preview data for import with column detection',
	)
	@PostMapping('/<int:id>/import/preview')
	@Authorized(True, permissions=['main.change_study'])
	def previewImportData(self, request, id: int):
		"""Preview import data - delegated to StudyService."""
		data = request.data
		file_url = data.get('fileUrl')
		mapping = data.get('mapping')

		if not file_url:
			return Return.badRequest('No file URL provided')

		try:
			result = self.studyService.previewDataImport(
				study_id=id,
				file_url=file_url,
				mapping=mapping,
			)
			return Return.ok(result)
		except ValueError as e:
			Logger.error(f'Error previewing import: {e}')
			return Return.badRequest(str(e))
		except Exception as e:
			Logger.error(f'Error previewing import: {e}')
			return Return.badRequest(f'Failed to parse file: {e}')

	@extend_schema(
		tags=['Dataset'],
		summary='Execute data import',
		description='Execute the data import from uploaded file into the dataset',
	)
	@PostMapping('/<int:id>/import/execute')
	@Authorized(True, permissions=['main.change_study'])
	def executeImportData(self, request, id: int):
		"""Execute data import - delegated to StudyService."""
		data = request.data
		file_url = data.get('fileUrl')
		mapping = data.get('mapping', {})
		column_types = data.get('columnTypes', {})

		if not file_url:
			return Return.badRequest('No file URL provided')

		try:
			result = self.studyService.executeDataImport(
				study_id=id,
				file_url=file_url,
				mapping=mapping,
				column_types=column_types,
				created_by=request.user,
			)
			return Return.ok(result)
		except ValueError as e:
			Logger.error(f'Error executing import: {e}')
			return Return.badRequest(str(e))
		except Exception as e:
			Logger.error(f'Error executing import: {e}')
			return Return.badRequest(f'Import failed: {e}')

	@extend_schema(
		tags=['Dataset'],
		summary='Execute data import with streaming progress',
		description='Execute the data import with real-time progress updates via Server-Sent Events',
	)
	@PostMapping('/<int:id>/import/execute-stream')
	@Authorized(True, permissions=['main.change_study'])
	def executeImportDataStream(self, request, id: int):
		"""Execute data import with streaming progress updates (SSE)."""
		import asyncio
		import json
		from concurrent.futures import ThreadPoolExecutor
		from queue import Empty, Queue

		from asgiref.sync import sync_to_async
		from django.http import StreamingHttpResponse

		Logger.info(f'[STREAM] Starting streaming import for dataset {id}')
		
		data = request.data
		file_url = data.get('fileUrl')
		mapping = data.get('mapping', {})
		column_types = data.get('columnTypes', {})

		Logger.info(f'[STREAM] file_url: {file_url}')
		Logger.info(f'[STREAM] mapping keys: {list(mapping.keys())}')

		if not file_url:
			Logger.error('[STREAM] No file URL provided')
			return Return.badRequest('No file URL provided')

		# Store request context for the generator
		user = request.user
		study_service = self.studyService

		# Use a queue to pass events from sync thread to async generator
		event_queue = Queue()

		def run_sync_import():
			"""Run the sync import in a separate thread and put events in queue."""
			try:
				for event in study_service.executeDataImportStream(
					study_id=id,
					file_url=file_url,
					mapping=mapping,
					column_types=column_types,
					created_by=user,
				):
					event_queue.put(event)
				event_queue.put(None)  # Signal completion
			except Exception as e:
				Logger.error(f'[STREAM] Sync import error: {e}')
				event_queue.put({'type': 'error', 'message': str(e)})
				event_queue.put(None)

		async def generate_events_async():
			"""Async generator that reads from queue populated by sync thread."""
			Logger.info('[STREAM] Async generator started')
			event_count = 0
			
			# Start sync import in thread pool
			loop = asyncio.get_event_loop()
			executor = ThreadPoolExecutor(max_workers=1)
			loop.run_in_executor(executor, run_sync_import)
			
			try:
				while True:
					# Poll queue with async sleep to not block
					try:
						event = event_queue.get_nowait()
					except Empty:
						await asyncio.sleep(0.05)  # Small delay between polls
						continue
					
					if event is None:  # Completion signal
						Logger.info(f'[STREAM] Async generator completed, yielded {event_count} events')
						break
					
					event_count += 1
					event_type = event.get('type', 'progress')
					sse_data = f'event: {event_type}\ndata: {json.dumps(event)}\n\n'
					Logger.info(f'[STREAM] Yielding event #{event_count}: type={event_type}, current={event.get("current", "N/A")}, total={event.get("total", "N/A")}')
					yield sse_data.encode('utf-8')
					
			except Exception as e:
				Logger.error(f'[STREAM] Streaming error: {e}')
				error_data = f'event: error\ndata: {json.dumps({"type": "error", "message": str(e)})}\n\n'
				yield error_data.encode('utf-8')
			finally:
				executor.shutdown(wait=False)

		Logger.info('[STREAM] Creating StreamingHttpResponse with async generator')
		response = StreamingHttpResponse(
			generate_events_async(),
			content_type='text/event-stream; charset=utf-8'
		)
		response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
		response['Pragma'] = 'no-cache'
		response['Expires'] = '0'
		response['X-Accel-Buffering'] = 'no'  # Disable nginx buffering
		response['Connection'] = 'keep-alive'
		Logger.info('[STREAM] Returning response')
		return response

